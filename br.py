import re
from openpyxl import Workbook,load_workbook
import os
import time
import datetime

LEN = 0
HEAD = ["日期","时间","金额","支付方式","用途"]
WAY={"WECHAT":("WECHAT","微信单号/微信转账单号"),"ALIPAY":("ALIPAY","支付宝订单号")}
re_date = re.compile(r"([0-9]{1,4})/0?([0-9]{1,2})/0?([0-9]{1,2})")
re_mon = re.compile(r"([0-9]{1,4})/0?([0-9]{1,2})")
global DATE,FILE, DATA, SHEET
'''global 
FILE 当前打开的文件
DATA
SHEET'''
def init():
    global FILE, DATA, SHEET
    _time = time.localtime()
    FILE = f"{_time.tm_year}年{_time.tm_mon}月.xlsx"
    initFile(True)

def initFile(isNeedCreate=False):
    global FILE,DATA,SHEET
    if not os.listdir().__contains__(FILE):
        if isNeedCreate:
            DATA = Workbook()
            end()     
        else:
            print("no workbook now!")
    else:
        DATA =  load_workbook(FILE)
    SHEET=DATA.active
    SHEET["A1"]="日期"
    SHEET["B1"]="时间"
    SHEET["C1"]="金额"
    SHEET["D1"]="支付方式"
    SHEET["E1"]="资金来源"
    SHEET["F1"]="用途"
    print("\t\t{}   OPEN Succeed!".format(FILE))

def end():
    DATA.save(filename=FILE)
'''globa
DATE 当前日期'''
def getTime(isNeedUpdate=False):
    global DATE
    _time = time.localtime()
    if isNeedUpdate:
        DATE = datetime.datetime(_time.tm_year,_time.tm_mon,_time.tm_mday)
    date = f"{_time.tm_year}/{_time.tm_mon}/{_time.tm_mday}" 
    mtime = f"{_time.tm_hour}:{_time.tm_min}:{_time.tm_sec}"
    return date,mtime

def menu():
    navs=["1.添加微信支付记录","2.添加支付宝支付记录","3.其他消费记录增加","4.查询当月消费记录","5.查询当日消费记录","6.自定义查询","7.自定义日期添加","8.回到当前时间"]
    for nav in navs:
        print("\t\t"+nav+"\n")

def main():
    init()
    getTime(True)
    menu()
    while True:
        choice = input("请输入编号:")
        if choice == "1":
            add (WAY["WECHAT"])
        elif choice == "2":
            add(WAY["ALIPAY"])
        elif choice == "3":
            way = input("请输入消费方式('平台，交易识别号'):")
            add(way.split('，'))
        elif choice == "4":  
            query(DATE)
        elif choice == "5":  
            query(DATE,condition="day")
        elif choice == "6":  
            customizeQuery()
        elif choice == "7":
            specificDateInsert()
        elif choice =="8":
            init()
            getTime(True)
        else:
            answer = input("error! exit?(y/n)")
            if answer.upper() == "Y":
                break
            elif answer.upper() == "N":
                continue
    end()

    
def modifyDateWithInput():
    global FILE, DATA, SHEET
    datein = input("输入日期或月份(yy/MM/dd or yy/MM)")
    if modifyDateWithArg(datein):
        return datein
    else:
        print("error! back to menu...")
        return False

def modifyDateWithArg(datein,isNeedCreate=False):
    global SHEET,FILE,DATA
    if re_date.fullmatch(datein) or re_mon.fullmatch(datein):
        FILE = "{}年{}月.xlsx".format(re_mon.match(datein)[1],re_mon.match(datein)[2])
        if not os.listdir().__contains__(FILE):
            if isNeedCreate:
                initFile(True)
                return
            else:
                print(f"{FILE}file not found error! back to menu...")
            return False
        DATA = load_workbook(FILE)
        print("\t\t{}   OPEN Succeed!".format(FILE))
        SHEET = DATA.active  
        return datein

def specificDateInsert():
    global DATE
    datein=modifyDateWithInput()
    if re_date.fullmatch(datein):
        date = re_date.fullmatch(datein)
        DATE = datetime.datetime(int (date[1]),int(date[2]),int(date[3]))
        print("now date:{}".format(DATE))
    else:
        print("not a day, back today");
        init()

def add(way):
    while True:
        time=getTime()[1]
        try:
            money = float(input("请输入金额:"))
        except:
            answer = input("error! exit?(y/n)")
            if answer.upper() == "Y":
                break
            elif answer.upper() == "N":
                continue
            else:
                print(answer.upper())
                print("error! back to menu...")
                return
        src = input("请输入资金来源(去路):")
        purpose = input("请输入用途:")
        id= way[1]+input("请输入{}:".format(way[1]))
        args=[DATE,time,money,way[0],src,purpose+"("+id+")"]
        insertOne(args)
        end()

def insertOne(args):
    SHEET.append(args)
def query(arg, condition = "mon"):
    total = 0
    rows = SHEET.rows
    if condition == "mon":
        for row in rows:
            print( "|{0}|{1:<10}|{2:<10}|{3:<10}|{4:<10}|{5:<50}".format(row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value) )
            if row[0].value != "日期":
                total = total + row[2].value
    elif condition == "day" : 
        for row in rows:
            if row[0].value == arg:
                print( "|{0}|{1:<10}|{2:<10}|{3:<10}|{4:<10}|{5:<50}".format(row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value) )
                total = total + row[2].value
            elif row[0].value == "日期":
                print( "|{0}|{1:<10}|{2:<10}|{3:<10}|{4:<10}|{5:<50}".format(row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value) )
              
    print(f"TOTAL:{total}")

def customizeQuery():
    datein=modifyDateWithInput()
    if datein:
        if re_date.fullmatch(datein):       
            date = re_date.fullmatch(datein)
            query(datetime.datetime(int (date[1]),int(date[2]),int(date[3])) ,condition="day")
        elif re_mon.fullmatch(datein):
            query(DATE,condition="mon") 
        init()
    else:
        return   

if __name__ == "__main__": 
    main()
    # init()
    # getTime()
    # query(DATE,condition="day")



